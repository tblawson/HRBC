# -*- coding: utf-8 -*-
"""
R_info.py

Created on Fri Sep 18 09:40:33 2015

@author: t.lawson
"""
import string
import datetime as dt
import time
import math
import xlrd
from openpyxl.styles import Font, PatternFill, Border, Side
import GTC
from numbers import Number

RL_SEARCH_LIMIT = 500

INF = 1e6  # 'inf' dof
ZERO = GTC.ureal(0, 0)

# Use for G1 & G2 in AUTO mode:
Vgain_codes_auto = {0.1: 'Vgain_0.1r0.1', 0.5: 'Vgain_0.5r1', 0.9: 'Vgain_1r1',
                    1.0: 'Vgain_1r1', 5.0: 'Vgain_5r10', 9.0: 'Vgain_10r10',
                    10.0: 'Vgain_10r10', 100.0: 'Vgain_100r100'}
# Use for G2 in FIXED mode:
Vgain_codes_fixed = {0.1: 'Vgain_0.5r1', 0.5: 'Vgain_0.5r1', 0.9: 'Vgain_1r1',
                     1.0: 'Vgain_1r10', 5.0: 'Vgain_5r10', 9.0: 'Vgain_10r10', 
                     10.0: 'Vgain_10r100', 100.0: 'Vgain_100r100'}


# ______________________________Useful funtions:______________________________

def Make_Log_Name(v):
    return 'HRBAv'+str(v)+'_'+str(dt.date.today())+'.log'


# Extract resistor names from comment
"""
Parse first part of comment for resistor names.
Names must appear immediately after the strings 'R1: ' and 'R2: ' and
immediately before the string ' monitored by GMH'.
"""


def ExtractNames(comment):
    assert comment.find('R1: ') >= 0, 'R1 name not found in comment!'
    assert comment.find('R2: ') >= 0, 'R2 name not found in comment!'
    R1_name = comment[comment.find('R1: ') + 4:comment.find(' monitored by GMH')]
    R2_name = comment[comment.find('R2: ') + 4:comment.rfind(' monitored by GMH')]
    return R1_name, R2_name


def GetRval(name):
    """
    Extract nominal resistor value from name.
    Parse the resistor name for the nominal value. Resistor names MUST be of the form
    'xxx nnp', where 'xxx ' is a one-word description ending with a SINGLE SPACE,
    'nn' is an integer (usually a decade value) and the last character 'p' is a letter
    indicating a decade multiplier.
    """
    prefixes = {'r': 1, 'R': 1, 'k': 1000, 'M': 1e6, 'G': 1e9}

    if name[-1] in prefixes:
        mult = prefixes[name[-1]]
    else:
        mult = 0
    assert mult != 0, 'Error parsing resistor name - unkown multiplier!'

    # return numeric part of last word, multiplied by 1, 10^3, 10^6 or 10^9:
    r_val_str = name.split()[-1]
    return mult * int(r_val_str.strip(string.ascii_letters))
    # return mult*int(string.strip(string.split(name)[-1], string.letters))


def GetRLstartrow(sheet, Id, jump, log):
    search_row = 1
    while search_row < RL_SEARCH_LIMIT:  # Don't search forever.
        result = sheet['A'+str(search_row)].value  # scan down column A
        if result == 'Run Id:':
            RL_Id = sheet['B'+str(search_row)].value  # Find a run Id
            if RL_Id == Id:  # Found the right data, so we're done.
                RL_start = search_row + 1  # 1st line of Rlink data's header (excl Id)
                return RL_start
            else:  # Jump to just before start of next data-block
                search_row += jump
        search_row += 1
    print('No matching Rlink data!')
    log.write('GetRLstartrow():No matching Rlink data!\n')
    return -1


# Convert list of data to ureal, where possible
def Uncertainize(row_items):
    v = row_items[2]
    u = row_items[3]
    d = row_items[4]
    l = row_items[5]
    if (u is not None) and isinstance(v, Number):
        if d == u'inf':
            un_num = GTC.ureal(v, u, label=l)  # default dof = inf
        else:
            un_num = GTC.ureal(v, u, d, l)
        return un_num
    else:  # non-numeric value
        return v


# Convert a resistive T-sensor reading from resistance to temperature
def R_to_T(alpha, beta, R, R0, T0):
    if beta == 0:  # no 2nd-order T-Co
        T = (R/R0 - 1)/alpha + T0
    else:
        a = beta
        b = alpha-2*T0
        c = 1-alpha*T0 + beta*T0**2 - (R/R0)
        T = (-b + GTC.sqrt(b**2-4*a*c))/(2*a)
    return T


# Return average of a list of time-strings("%d/%m/%Y %H:%M:%S")
# as a time string or float
def av_t_strin(t_list, switch):
    assert switch in ('fl', 'str'), 'Unknown switch for function av_t_strin()!'
    throwaway = dt.datetime.strptime('20110101', '%Y%m%d')  # known bug fix
    n = float(len(t_list))
    t_av = 0.0
    for s in t_list:
        """
        Excel cells may contain an un-formatted time as a float,
        or an Excel date-time format string:
        """
        if isinstance(s, str):  # Python2.7: type(s) is 'unicode'
            t_dt = dt.datetime.strptime(s, '%d/%m/%Y %H:%M:%S')
        elif isinstance(s, float):
            print(s, 'is a float...')
            t_dt = xlrd.xldate.xldate_as_datetime(s, 0)
        elif isinstance(s, dt.datetime):
            t_dt = s
        else:
            assert 0, f'Time format is not unicode or float! - {s}'
        t_tup = dt.datetime.timetuple(t_dt)
        t_av += time.mktime(t_tup)

    t_av /= n  # av. time as float (seconds from epoch)
    if switch == 'fl':
        return t_av
    elif switch == 'str':
        t_av_fl = dt.datetime.fromtimestamp(t_av)
        return t_av_fl.strftime('%d/%m/%Y %H:%M:%S')  # av. time as string


# Write headings on Summary sheet
def WriteHeadings(sheet, row, version, start_row, end_row):
    now = dt.datetime.now()
    sheet['A' + str(row-1)].font = Font(b=True)
    sheet['A' + str(row-1)] = f'Data-rows {start_row} -> {end_row}, processed with HRBA v'\
                            +str(version)+' on '+now.strftime("%A, %d. %B %Y %I:%M%p")
    sheet['J' + str(row)] = 'Uncertainty Budget'

    sheet['R' + str(row)] = 'R1(T)'
    sheet['U' + str(row)] = 'exp. U(95%)'
    sheet['V' + str(row)] = 'av T'
    sheet['Y' + str(row)] = 'av date/time'
    sheet['Z' + str(row)] = 'av V'
    sheet['AC' + str(row)] = 'k'
    sheet['AC' + str(row - 1)] = 'T'
    sheet['AD' + str(row)] = 'exp. U(95%)'
    sheet['AE' + str(row)] = 'k'
    sheet['AE' + str(row - 1)] = 'V'
    sheet['AF' + str(row)] = 'exp. U(95%)'
    sheet['AG' + str(row)] = 'val'
    sheet['AG' + str(row - 1)] = '%RH'
    sheet['AH' + str(row)] = 'sd'
    sheet['AI' + str(row)] = 'dof'
    sheet['AJ' + str(row)] = 'k'
    sheet['AK' + str(row)] = 'exp. U(95%)'
    sheet['AL' + str(row)] = 'CMC (ppm)'
    sheet['AM' + str(row)] = 'CMC (Ohm)'

    row += 1
    sheet['A' + str(row)] = 'Name'
    sheet['B' + str(row)] = 'Test V'
    sheet['C' + str(row)] = 'Date'
    sheet['D' + str(row)] = 'T'
    sheet['E' + str(row)] = 'R1'
    sheet['F' + str(row)] = 'std u.'
    sheet['G' + str(row)] = 'dof'
    sheet['H' + str(row)] = 'exp. U(95%)'
    sheet['J' + str(row)] = 'Quantity (Label)'
    sheet['K' + str(row)] = 'Value'
    sheet['L' + str(row)] = 'Std u'
    sheet['M' + str(row)] = 'dof'
    sheet['N' + str(row)] = 'sens. coef.'
    sheet['O' + str(row)] = 'u contrib.'

    sheet['Q'+str(row)] = 'LV'
    row += 1
    sheet['Q'+str(row)] = 'HV'

    return row


# Write measurement summary   
def WriteThisResult(sheet, row, result):
    # print(f"\tWriteThisResult(): RHs: {result['RHs']}")
    sheet['A'+str(row)].font = Font(color='FFFF00')  # YELLOW
    sheet['A'+str(row)].fill = PatternFill(patternType='solid',
                                           fgColor='FF0000')  # RED
    sheet['A'+str(row)] = str(result['name'])

    sheet['B'+str(row)] = result['V'].x
    sheet['B'+str(row+1)] = result['V'].u
    sheet['B'+str(row+2)] = result['V'].df

    sheet['C'+str(row)] = str(result['time_str'])

    temperature = result['T1_A'] + result['Tdef1']
    sheet['D'+str(row)] = temperature.x  # result['T']
    sheet['D'+str(row+1)] = temperature.u  # result['T']
    sheet['D'+str(row+2)] = temperature.df  # result['T']

    sheet['E'+str(row)] = result['R'].x

    sheet['F'+str(row)] = result['R'].u

    if math.isinf(result['R'].df):
        sheet['G'+str(row)] = str(result['R'].df)
    else:
        sheet['G'+str(row)] = round(result['R'].df)

    # Exp Uncert:
    sheet['H'+str(row)] = result['R_expU']


# Sorting helper function - sort by uncert. contribution
def by_u_cont(line):
    return line[5]


def WriteBudget(sheet, row, budget):
    for line in budget:
        sheet['J'+str(row)] = line[0]  # Quantity (label)
        sheet['K'+str(row)] = line[1]  # Value
        sheet['L'+str(row)] = line[2]  # Uncert.
        if math.isinf(line[3]):
            sheet['M'+str(row)] = str(line[3])  # dof
        else:
            sheet['M'+str(row)] = round(line[3])  # dof
        sheet['N'+str(row)] = line[4]  # Sens. coef.
        sheet['O'+str(row)] = line[5]  # Uncert. contrib.
        row += 1
    return row


def add_if_unique(item, lst):
    """
    Append 'item' to 'lst' only if it is not already present.
    """
    if item not in lst:
        lst.append(item)
    return lst


# Weighted least-squares fit (R1-T)
def write_R1_T_fit(results, RH_data, sheet, row, log, Tdef, RH_def, RH_cor, R1_alpha):
    T_data = [T for T in [result['T1_A'] for result in results]]  # All T ureals
    # RH_data = []
    # for result in results:
    #     # print(f"'{result['time_str']} - {result['RHs']}'")
    #     RH_data.extend(result['RHs'])

    unique_T_data = []
    for T in T_data:
        add_if_unique(T, unique_T_data)

    T_av = GTC.fn.mean(T_data) + Tdef  # Type-B added here. ONCE!
    # RH_data are plain numbers so use ta.estimate:
    RH_av = GTC.ta.estimate(RH_data) + RH_def + RH_cor  # Type-Bs added here. ONCE!
    T_rel = [t_k - T_av for t_k in T_data]  # x-vals
    alpha = GTC.ureal(0, 0)  # Pre-defined default - will be updated later.

    y = [R for R in [result['R'] for result in results]]  # All R values
    u_y = [R.u for R in [result['R'] for result in results]]  # All R uncerts

    if len(unique_T_data) <= 1:  # No T-variation recorded, so can't fit to T. CHANGED from len(set(T_data)).
        R1 = GTC.fn.mean(y)
        print('R1_LV (av, not fit):', R1)
        log.write('\nR1_LV (av, not fit): ' + str(R1))
    else:
        # a_ta,b_ta = GTC.ta.line_fit_wls(T_rel,y,u_y).a_b
        # Assume uncert of individual measurements dominate uncert of fit
        R1, alpha = GTC.ta.line_fit_wls(T_rel, y, u_y).a_b
        print(f'Fit params:\t intercept={R1.x}+/-{R1.u},dof={R1.df}. Slope={alpha.x}+/-{alpha.u},dof={alpha.df}')
        log.write(f'Fit params:\t intercept={R1.x}+/-{R1.u},dof={R1.df}. Slope={alpha.x}+/-{alpha.u},dof={alpha.df}')

    T_def_duc = GTC.ureal(0, T_av.u, T_av.df, label='T_def_duc')
    T1_def_on_R1 = GTC.fn.mul2(R1_alpha, T_def_duc)
    R1 = R1*(1 + T1_def_on_R1)  # Include Tdef1 influence on final DUC value here.

# R1 measurement:
    sheet['R'+str(row)] = R1.x
    sheet['S'+str(row)] = R1.u
    if math.isinf(R1.df):
        sheet['T'+str(row)] = str(R1.df)
    else:
        sheet['T'+str(row)] = round(R1.df)

    sheet['U'+str(row)] = R1.u*GTC.rp.k_factor(R1.df)
# T1 measurement:
    sheet['V'+str(row)] = T_av.x
    sheet['W'+str(row)] = T_av.u
    if math.isinf(T_av.df):
        sheet['X'+str(row)] = str(T_av.df)  # 'inf'
    else:
        sheet['X'+str(row)] = round(T_av.df, 1)
    T_av_k = GTC.rp.k_factor(T_av.df)
    sheet['AC'+str(row)] = T_av_k
    sheet['AD'+str(row)] = T_av_k*T_av.u

# date / time:
    t = [result['time_fl'] for result in results]  # x data (time,s from epoch)
    t_av = GTC.ta.estimate(t)
    time_av = dt.datetime.fromtimestamp(t_av.x)  # A Python datetime object
    sheet['Y'+str(row)] = time_av.strftime('%d/%m/%Y %H:%M:%S')  # string-formatted for display

# V1 measurement:
    V1 = [V for V in [result['V'] for result in results]]
    V_av = GTC.fn.mean(V1)
    sheet['Z'+str(row)] = V_av.x
    sheet['AA'+str(row)] = V_av.u
    if math.isinf(V_av.df):
        sheet['AB'+str(row)] = str(V_av.df)
    else:
        sheet['AB'+str(row)] = round(V_av.df)
    V_av_k = GTC.rp.k_factor(V_av.df)
    sheet['AE' + str(row)] = V_av_k
    sheet['AF' + str(row)] = V_av_k * V_av.u

# %RH measurement:
    sheet['AG' + str(row)] = RH_av.x
    sheet['AH' + str(row)] = RH_av.u
    if math.isinf(RH_av.df):
        sheet['AI' + str(row)] = str(RH_av.df)  # 'inf'
    else:
        sheet['AI' + str(row)] = round(RH_av.df, 1)
    RH_av_k = GTC.rp.k_factor(RH_av.df)
    sheet['AJ' + str(row)] = RH_av_k
    sheet['AK' + str(row)] = RH_av_k * RH_av.u

# CMCs:
    CMC_ppm = 0.7+27*R1.x/1e9 - 20*GTC.pow(R1.x/1e9, 3)
    sheet['AL' + str(row)] = CMC_ppm
    sheet['AM' + str(row)] = R1.x*CMC_ppm/1e6

    return R1, alpha, T_av, V_av, time_av


def create_comment_ref(version):
    now_tup = dt.datetime.now()
    now_fmt = now_tup.strftime('%d/%m/%Y %H:%M:%S')
    return 'HRBA' + str(version) + '_' + now_fmt


def create_label(name, param, Id):
    return name.split()[0] + '_' + param + '_' + Id


def find_param_row(sheet, max_row, name, param):
    row_num = 0
    for r in sheet.iter_rows(max_col=2, max_row=max_row, values_only=True):
        row_num += 1
        if r[0] == name and r[1] == param:
            break
    return row_num


def update_R_Info(sheet, param, value, row, label, version):
    """
    Update the parameters of a known resistor.
    :param sheet: openpyxl sheet object for 'Parameters' sheet
    :param param: parameter (str)
    :param value: data (usually a GTC.ureal)
    :param row: current row of 'Parameters' sheet (int)
    :param label: ureal label (str)
    :param version: HRBA version (str)
    :return: none
    """

    # Don't update alpha (calculated estimate is usually unreliable).
    # Also, exclude non-GTC.ureal params
    # if param not in ('date', 'T_sensor', 'alpha'):
    if param == 'date':
        sheet['C' + str(row)] = value
    else:
        sheet['C' + str(row)] = value.x
        sheet['D' + str(row)] = value.u
        if math.isinf(value.df):
            sheet['E' + str(row)] = str(value.df)
        else:
            sheet['E' + str(row)] = round(value.df)
            sheet['F' + str(row)] = label
        sheet['G' + str(row)] = create_comment_ref(version)


def append_R_Info(name, params, data, sheet, row, Id, v):
    """
    Adds a new resistor entry to the Parameters sheet, if it was
    previously not present.
    :param name: resistor name (str)
    :param params: list of parameters (list of str)
    :param data: list of values (list of various types, including ureal)
    :param sheet: Openpyxl sheet object for 'Parameters' sheet
    :param row: current row in 'Parameters' sheet to start writing (int)
    :param Id: Run ID - used to construct ureal label (str)
    :param v: HRBA version - used to construct comment/reference (str)
    :return: current row for appending mor data to the sheet (int)
    """
    R_dict = dict(zip(params, data))

    for param in params:
        label = create_label(name, param, Id)
        row += 1
        sheet['A'+str(row)] = name
        sheet['B'+str(row)] = param
        update_R_Info(sheet, param, R_dict[param].u, row, label, v)

    # Mark end of data with a bottom border on cells of last row:
    b = Border(bottom=Side(style='thin'))
    sheet['A'+str(row)].border = b
    sheet['B'+str(row)].border = b
    sheet['C'+str(row)].border = b
    sheet['D'+str(row)].border = b
    sheet['E'+str(row)].border = b
    sheet['F'+str(row)].border = b
    sheet['G'+str(row)].border = b

    return row


def GetDigi(readings):
    """
    Return maximum digitization level of a set of data.
    """
    max_n = 0
    for reading in readings:
        # if 'e' in str(reading) or 'E' in str(reading):
        #     (m, e) = math.modf(reading)
        #     n = int(str(m).split('e')[1])
        # else:
        n = -1*len(str(float(reading)).split('.')[1])

        if abs(n) > abs(max_n):
            max_n = n
    d = 10**max_n
    print(f'GetDigi(): Max digitisation is {d}')
    return d

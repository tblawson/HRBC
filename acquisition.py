
# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 09:36:42 2015

WORKING VERSION

@author: t.lawson


acquisition.py:
Thread class that executes processing.
Contains definitions for usual __init__() and run() methods
 AND an abort() method. The Run() method forms the core of the
 procedure - any changes to the way the measurements are taken
 should be made here, and within included subroutines.
"""
import wx
from threading import Thread
import datetime as dt
import time

import numpy as np

from openpyxl.styles import Font, Border, Side, colors

import HighRes_events as evts
import devices  # visastuff
# import devices as GMH


class AqnThread(Thread):
    """Acquisition Thread Class."""
    def __init__(self, parent):
        # This runs when an instance of the class is created
        Thread.__init__(self)
        self.RunPage = parent
        self.SetupPage = self.RunPage.GetParent().GetPage(0)
        self.PlotPage = self.RunPage.GetParent().GetPage(2)
        self.TopLevel = self.RunPage.GetTopLevelParent()
        self.Comment = self.RunPage.Comment.GetValue()
        self._want_abort = 0

        self.V1Data = []
        self.V2Data = []
        self.VdData = []
        self.V1Times = []
        self.V2Times = []
        self.VdTimes = []

        self.V1_set = 0.0
        self.V2_set = 0.0
        self.start_del = 0.0
        self.n_readings = 0
        self.AZ1_del = 0.0
        self.range_del = 0.0
        self.T1 = 0.0
        self.T2 = 0.0
        self.Troom = 0.0
        self.Proom = 0.0
        self.RHroom = 0.0

        self.log = self.SetupPage.log

        self.Range_Mode = {True: 'AUTO', False: 'FIXED'}

        print('Role -> Instrument:')
        print('Role -> Instrument:', file=self.log)
        print('------------------------------')
        print('------------------------------', file=self.log)

        # Print all GPIB instrument objects
        for r in devices.ROLES_WIDGETS.keys():
            d = devices.ROLES_WIDGETS[r]['icb'].GetValue()
            # For 'switchbox' role, d is actually the setting
            # (V1, Vd1,...) not the instrument description.

            print('{} -> {}'.format(devices.INSTR_DATA[d]['role'], d))
            print('{} -> {}'.format(devices.INSTR_DATA[d]['role'], d), file=self.log)
            if r != devices.INSTR_DATA[d]['role']:
                devices.INSTR_DATA[d]['role'] = r
                print('Role data corrected to: {} -> {}'.format(r, d))
                print('Role data corrected to: {} -> {}'.format(r, d), file=self.log)

        # Get filename of Excel file
        self.xlfilename = self.SetupPage.XLFile.GetValue()  # Full path
        self.path_components = self.xlfilename.split('\\')  # List of all the bits between '\'s
        self.directory = '\\'.join(self.path_components[0:-1])

        # Find existing workbook
        self.wb_io = self.SetupPage.wb
        self.ws = self.wb_io.get_sheet_by_name('Data')

        # read start/stop row numbers from Excel file
        self.start_row = self.ws['B1'].value
        self.stop_row = self.ws['B2'].value
        strt_ev = evts.StartRowEvent(row=self.start_row)
        wx.PostEvent(self.RunPage, strt_ev)
        stp_ev = evts.StopRowEvent(row=self.stop_row)
        wx.PostEvent(self.RunPage, stp_ev)

        self.settle_time = self.RunPage.SettleDel.GetValue()

        # Local record of GMH ports and addresses
        self.GMH1Demo_status = devices.ROLES_INSTR['GMH1'].demo
        self.GMH2Demo_status = devices.ROLES_INSTR['GMH2'].demo
        self.GMH1Port = devices.ROLES_INSTR['GMH1'].addr
        self.GMH2Port = devices.ROLES_INSTR['GMH2'].addr

        self.start()  # Starts the thread running on creation

    def run(self):
        """
        Run Worker Thread.
        This is where all the important stuff goes, in a repeated cycle.
        """
        # Set button availability
        self.RunPage.StopBtn.Enable(True)
        self.RunPage.StartBtn.Enable(False)
        self.RunPage.RLinkBtn.Enable(False)

        # Clear plots
        clr_plot_ev = evts.ClearPlotEvent()
        wx.PostEvent(self.PlotPage, clr_plot_ev)

        # Column headings
        head_row = self.start_row-2  # Main headings
        sub_row = self.start_row-1  # Sub-headings
        # Write unique id for this run
        # - used to pair measurement data with RLink data
        self.ws['A'+str(sub_row)] = 'Run Id:'
        self.ws['B'+str(sub_row)].font = Font(b=True)
        self.ws['B'+str(sub_row)] = str(self.RunPage.run_id)
        self.ws['A'+str(head_row)] = 'V1_set'
        self.ws['B'+str(head_row)] = 'V2_set'
        self.ws['C'+str(head_row)] = 'n'
        self.ws['D'+str(head_row)] = 'Start/xl del.'
        self.ws['E'+str(head_row)] = 'AZ1 del.'
        self.ws['F'+str(head_row)] = 'Range del.'
        self.ws['G'+str(head_row)] = 'V2'
        self.ws['G'+str(sub_row)] = 't'
        self.ws['H'+str(sub_row)] = 'V'
        self.ws['I'+str(sub_row)] = 'sd(V)'
        # miss columns j,k,l
        self.ws['M'+str(head_row)] = 'Vd1'
        self.ws['M'+str(sub_row)] = 't'
        self.ws['N'+str(sub_row)] = 'V'
        self.ws['O'+str(sub_row)] = 'sd(V)'
        self.ws['P'+str(head_row)] = 'V1'
        self.ws['P'+str(sub_row)] = 't'
        self.ws['Q'+str(sub_row)] = 'V'
        self.ws['R'+str(sub_row)] = 'sd(V)'
        self.ws['S'+str(head_row)] = 'dvm_T1'
        self.ws['T'+str(head_row)] = 'dvm_T2'
        self.ws['U'+str(head_row)] = 'GMH_T1'
        self.ws['V'+str(head_row)] = 'GMH_T2'
        self.ws['W'+str(head_row)] = 'Ambient Conditions'
        self.ws['W'+str(sub_row)] = 'T'
        self.ws['X'+str(sub_row)] = 'P(mbar)'
        self.ws['Y'+str(sub_row)] = '%RH'
        self.ws['Z'+str(head_row)] = 'Comment'
        self.ws['AC'+str(head_row)] = 'Role'
        self.ws['AD'+str(head_row)] = 'Instrument descr.'
        self.ws['AE'+str(head_row)] = 'Range mode'

        stat_ev = evts.StatusEvent(msg='AqnThread.run():', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='Waiting to settle...', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

        time.sleep(self.settle_time)

        # Initialise all instruments (doesn't open GMH sensors yet)
        self.initialise()

        # write to both status fields:
        stat_ev = evts.StatusEvent(msg='', field='b')
        wx.PostEvent(self.TopLevel, stat_ev)

        stat_ev = evts.StatusEvent(msg='Post-initialise delay...', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)
        time.sleep(3)

        # Get some initial temperatures...
        T = devices.ROLES_INSTR['GMH1'].measure('T')
        self.ws['U'+str(self.start_row-1)] = T
        T = devices.ROLES_INSTR['GMH2'].measure('T')
        self.ws['V'+str(self.start_row-1)] = T

        # Record ALL roles and corresponding instr descriptions in XL sheet
        role_row = self.start_row
        bord_tl = Border(top=Side(style='thin'), left=Side(style='thin'))
        bord_tr = Border(top=Side(style='thin'), right=Side(style='thin'))
        bord_l = Border(left=Side(style='thin'))
        bord_r = Border(right=Side(style='thin'))
        bord_bl = Border(bottom=Side(style='thin'), left=Side(style='thin'))
        bord_br = Border(bottom=Side(style='thin'), right=Side(style='thin'))
        for r in devices.ROLES_WIDGETS.keys():
            if role_row == self.start_row:  # 1st row
                self.ws['AC'+str(role_row)].border = bord_tl
                self.ws['AD'+str(role_row)].border = bord_tr
            elif role_row == self.start_row + 9:  # last row
                self.ws['AC'+str(role_row)].border = bord_bl
                self.ws['AD'+str(role_row)].border = bord_br
            else:  # in-between rows
                self.ws['AC'+str(role_row)].border = bord_l
                self.ws['AD'+str(role_row)].border = bord_r
            self.ws['AC'+str(role_row)] = r
            d = devices.ROLES_WIDGETS[r]['icb'].GetValue()  # descr
            self.ws['AD'+str(role_row)] = d
            if r == 'DVM12':
                mode = self.Range_Mode[self.RunPage.RangeTBtn.GetValue()]
                self.ws['AE'+str(role_row)] = mode
            role_row += 1

        row = self.start_row
        pbar = 1

        # loop over xl rows..
        while row <= self.stop_row:
            if self._want_abort:
                self.abort_run()
                return

            if self._want_abort:
                self.abort_run()
                return
            stat_ev = evts.StatusEvent(msg='AqnThread.run():', field=0)
            wx.PostEvent(self.TopLevel, stat_ev)

            stat_ev = evts.StatusEvent(msg='Short delay 1...', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            time.sleep(5)

            self.set_up_meas_this_row(row)

            row_ev = evts.RowEvent(r=row)
            wx.PostEvent(self.RunPage, row_ev)

            #  V1...
            devices.ROLES_INSTR['DVM12'].send_cmd('LFREQ LINE')
            time.sleep(0.5)
            devices.ROLES_INSTR['DVM12'].send_cmd('DCV,' + str(int(self.V1_set)))
            if self._want_abort:
                self.abort_run()
                return

            stat_ev = evts.StatusEvent(msg='AqnThread.run():', field=0)
            wx.PostEvent(self.TopLevel, stat_ev)
            stat_ev = evts.StatusEvent(msg='Short delay 2...', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            time.sleep(3)

            # Set RS232 to V1
            cmd = devices.SWITCH_CONFIGS['V1']
            devices.ROLES_INSTR['switchbox'].send_cmd(cmd)
            self.SetupPage.Switchbox.SetValue('V1')  # update sw-box config icb
            devices.ROLES_INSTR['DVM12'].send_cmd('AZERO ON')
            if self._want_abort:
                self.abort_run()
                return

            stat_ev = evts.StatusEvent(msg='Range delay...', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            time.sleep(self.range_del)

            stat_ev = evts.StatusEvent(msg='Measuring V1', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            devices.ROLES_INSTR['DVM12'].read()  # junk = ...dvmV1V2
            devices.ROLES_INSTR['DVM12'].read()  # junk = ...dvmV1V2
            for i in range(self.n_readings):
                self.MeasureV('V1')
            self.T1 = devices.ROLES_INSTR['GMH1'].measure('T')

            # Update run displays on Run page via a DataEvent:
            av_t = np.mean(self.V1Times)[0]
            t1 = dt.datetime.fromtimestamp(av_t).strftime("%d/%m/%Y %H:%M:%S")
            V1m = np.mean(self.V1Data)[0]
            print('AqnThread.run(): V1m =', V1m)
            print('AqnThread.run(): V1m =', V1m, file=self.log)
            assert len(self.V1Data) > 1, "Can't take SD of one or less items!"
            V1sd = np.std(self.V1Data, ddof=1)[0]
            prog = 100.0 * pbar / (1 + self.stop_row - self.start_row)  # % progress
            update_ev = evts.DataEvent(t=t1, Vm=V1m, Vsd=V1sd, P=prog,
                                       r=row, flag='1')
            wx.PostEvent(self.RunPage, update_ev)

            #  V2...
            # Set RS232 to V2 BEFORE changing DVM range
            cmd = devices.SWITCH_CONFIGS['V2']
            devices.ROLES_INSTR['switchbox'].send_cmd(cmd)
            self.SetupPage.Switchbox.SetValue('V2')  # update sw-box config icb

            # If running with fixed range set range to 'str(self.V1_set)':
            if self.RunPage.RangeTBtn.GetValue():  # is True
                range2 = self.V2_set
            else:
                range2 = self.V1_set
            cmd = 'DCV,' + str(range2)
            devices.ROLES_INSTR['DVM12'].send_cmd(cmd)  # Reset DVM range
            if self._want_abort:
                self.abort_run()
                return
            time.sleep(0.5)  # was 0.1
            devices.ROLES_INSTR['DVM12'].send_cmd('LFREQ LINE')

            stat_ev = evts.StatusEvent(msg='AqnThread.run():', field=0)
            wx.PostEvent(self.TopLevel, stat_ev)
            stat_ev = evts.StatusEvent(msg='Short delay 3...', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            if self._want_abort:
                self.abort_run()
                return
            time.sleep(3)

            stat_ev = evts.StatusEvent(msg='Range delay...', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            if self._want_abort:
                self.abort_run()
                return
            time.sleep(self.range_del)

            stat_ev = evts.StatusEvent(msg='Measuring V2', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)

            devices.ROLES_INSTR['DVM12'].read()
            devices.ROLES_INSTR['DVM12'].read()
            for i in range(self.n_readings):
                self.MeasureV('V2')
            self.T2 = devices.ROLES_INSTR['GMH2'].measure('T')

            # Update displays on Run page via a DataEvent:
            av_t = np.mean(self.V2Times)[0]
            t2 = dt.datetime.fromtimestamp(av_t).strftime("%d/%m/%Y %H:%M:%S")
            V2m = np.mean(self.V2Data)[0]
            print('AqnThread.run(): V2m =', V2m)
            print('AqnThread.run(): V2m =', V2m, file=self.log)
            assert len(self.V2Data) > 1, "Can't take SD of one or less items!"
            V2sd = np.std(self.V2Data, ddof=1)[0]
            prog = 100.0 * pbar / (1 + self.stop_row - self.start_row)  # % progress
            update_ev = evts.DataEvent(t=t2, Vm=V2m, Vsd=V2sd, P=prog,
                                       r=row, flag='2')
            wx.PostEvent(self.RunPage, update_ev)

            #  Vd...
            # Set RS232 to Vd1
            cmd = devices.SWITCH_CONFIGS['Vd1']
            devices.ROLES_INSTR['switchbox'].send_cmd(cmd)
            self.SetupPage.Switchbox.SetValue('Vd1')  # update switchbox icb
            devices.ROLES_INSTR['DVMd'].send_cmd('RANGE AUTO')
            if self._want_abort:
                self.abort_run()
                return
            stat_ev = evts.StatusEvent(msg='Range delay...', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            time.sleep(self.range_del)

            stat_ev = evts.StatusEvent(msg='Measuring Vd', field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            devices.ROLES_INSTR['DVMd'].send_cmd('LFREQ LINE')
            devices.ROLES_INSTR['DVMd'].read()  # dummy read
            for i in range(self.n_readings):
                self.MeasureV('Vd')
            # Update displays on Run page via a DataEvent:
            av_t = np.mean(self.VdTimes)[0]
            td = dt.datetime.fromtimestamp(av_t).strftime("%d/%m/%Y %H:%M:%S")
            Vdm = np.mean(self.VdData)[0]
            print('AqnThread.run(): Vdm =', Vdm)
            print('AqnThread.run(): Vdm =', Vdm, file=self.log)
            assert len(self.VdData) > 1, "Can't take SD of one or less items!"
            Vdsd = np.std(self.VdData, ddof=1)[0]
            prog = 100.0 * pbar / (1 + self.stop_row - self.start_row)  # % progress

            # Record room conditions
            self.Troom = devices.ROLES_INSTR['GMHroom'].measure('T')
            self.Proom = devices.ROLES_INSTR['GMHroom'].measure('P')
            self.RHroom = devices.ROLES_INSTR['GMHroom'].measure('RH')

            self.write_data_this_row(row)

            # Plot data
            VdDates = []
            V1Dates = []
            V2Dates = []
            for d in self.VdTimes:
                VdDates.append(dt.datetime.fromtimestamp(d))
            for d in self.V1Times:
                V1Dates.append(dt.datetime.fromtimestamp(d))
            for d in self.V2Times:
                V2Dates.append(dt.datetime.fromtimestamp(d))
            clear_plot = 0
            if row == self.start_row:
                clear_plot = 1  # start each run with a clear plot
            plot_ev = evts.PlotEvent(td=VdDates, t1=V1Dates, t2=V2Dates,
                                     Vd=self.VdData, V1=self.V1Data,
                                     V2=self.V2Data, clear=clear_plot)
            wx.PostEvent(self.PlotPage, plot_ev)

            update_ev = evts.DataEvent(t=td, Vm=Vdm, Vsd=Vdsd, P=prog,
                                       r=row, flag='d')
            wx.PostEvent(self.RunPage, update_ev)            

            pbar += 1
            row += 1

        # (end of while loop):
        self.finish_run()
        return

    def initialise(self):
        # This is a Dascon (%RH) PLACEHOLDER for now...
        # Set Dascon Outlets 1,3 to 'On' and initialise (room T & RH)

        stat_ev = evts.StatusEvent(msg='Initialising instruments...', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)

        for r in devices.ROLES_INSTR.keys():
            d = devices.ROLES_WIDGETS[r]['icb'].GetValue()

            # Open non-GMH devices:
            if 'GMH' not in devices.ROLES_INSTR[r].Descr:
                print('AqnThread.initialise(): Opening {}.'.format(d))
                print('AqnThread.initialise(): Opening {}.'.format(d), file=self.log)
                devices.ROLES_INSTR[r].open()
            else:
                print('AqnThread.initialise(): {} - no initialisation needed.'.format(d))
                print('AqnThread.initialise(): {} - no initialisation needed.'.format(d), file=self.log)

            stat_ev = evts.StatusEvent(msg=d, field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            devices.ROLES_INSTR[r].init()
            time.sleep(1)
        stat_ev = evts.StatusEvent(msg='Done', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)

    def set_up_meas_this_row(self, row):
        d = devices.ROLES_INSTR['SRC2'].Descr
        if d.endswith('F5520A'):
            err = devices.ROLES_INSTR['SRC2'].check_err()
            print('Cleared F5520A error:', err)
            print('Cleared F5520A error:', err, file=self.log)
            time.sleep(3)  # Wait 3 s after checking error

        # Get V1,V2 setting, n, delays from spreadsheet
        self.V1_set = self.ws.cell(row=row, column=1).value
        self.RunPage.V1Setting.SetValue(str(self.V1_set))
        if self._want_abort:
                self.abort_run()
                return
        time.sleep(5)  # wait 5 s after setting voltage
        self.V2_set = self.ws.cell(row=row, column=2).value
        self.RunPage.V2Setting.SetValue(str(self.V2_set))
        self.start_del = self.ws.cell(row=row, column=4).value
        if self._want_abort:
                self.abort_run()
                return
        time.sleep(self.start_del)
        self.n_readings = self.ws.cell(row=row, column=3).value
        self.AZ1_del = self.ws.cell(row=row, column=5).value
        self.range_del = self.ws.cell(row=row, column=6).value
        del_ev = evts.DelaysEvent(n=self.n_readings,
                                  s=self.start_del,
                                  AZ1=self.AZ1_del,
                                  r=self.range_del)
        wx.PostEvent(self.RunPage, del_ev)

        del self.V1Data[:]
        del self.V2Data[:]
        del self.VdData[:]
        del self.V1Times[:]
        del self.V2Times[:]
        del self.VdTimes[:]

    def MeasureV(self, node):
        assert node in ('V1', 'V2', 'Vd'), 'Unknown argument to MeasureV().'
        if node == 'V1':
            self.V1Times.append(time.time())
            if devices.ROLES_INSTR['DVM12'].demo is True:
                dvm_op = np.random.normal(self.V1_set, 1.0e-5*abs(self.V1_set))
                self.V1Data.append(dvm_op)
            else:
                # lfreq line, azero once,range auto, wait for settle
                dvm_op = devices.ROLES_INSTR['DVM12'].read()
                self.V1Data.append(float(filter(self.filt, dvm_op)))
        elif node == 'V2':
            self.V2Times.append(time.time())
            if devices.ROLES_INSTR['DVM12'].demo is True:
                dvm_op = np.random.normal(self.V2_set, 1.0e-5*abs(self.V2_set))
                self.V2Data.append(dvm_op)
            else:
                dvm_op = devices.ROLES_INSTR['DVM12'].read()
                self.V2Data.append(float(filter(self.filt, dvm_op)))
        elif node == 'Vd':
            self.VdTimes.append(time.time())
            if self.AZ1_del > 0:
                devices.ROLES_INSTR['DVMd'].send_cmd('AZERO ONCE')
                time.sleep(self.AZ1_del)
            if devices.ROLES_INSTR['DVMd'].demo is True:
                dvm_op = np.random.normal(0.0, 1.0e-6)
                self.VdData.append(dvm_op)
            else:
                dvm_op = devices.ROLES_INSTR['DVMd'].read()
                self.VdData.append(float(filter(self.filt, dvm_op)))
            return 1

    def write_data_this_row(self, row):
        m = 'AqnThread.WriteDataThisRow():'
        stat_ev = evts.StatusEvent(msg=m, field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='Row '+str(row), field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

        av_t = np.mean(self.V1Times)[0]
        fmt = "%d/%m/%Y %H:%M:%S"
        t_stamp = str(dt.datetime.fromtimestamp(av_t).strftime(fmt))
        self.ws['P'+str(row)] = t_stamp
        print('WriteDataThisRow(): cell P{}: {}'.format(str(row), t_stamp), file=self.log)

        V1_mean = np.mean(self.V1Data)[0]
        self.ws['Q'+str(row)] = V1_mean
        print('WriteDataThisRow(): cell Q{}: {}'.format(str(row), V1_mean), file=self.log)

        V1_sd = np.std(self.V1Data, ddof=1)[0]
        self.ws['R'+str(row)] = V1_sd
        print('WriteDataThisRow(): cell R{}: {}'.format(str(row), V1_sd), file=self.log)

        av_t = np.mean(self.V2Times)[0]
        t_stamp = str(dt.datetime.fromtimestamp(av_t).strftime(fmt))
        self.ws['G'+str(row)] = t_stamp
        print('WriteDataThisRow(): cell G{}: {}'.format(str(row), t_stamp), file=self.log)

        V2_mean = np.mean(self.V2Data)[0]
        self.ws['H'+str(row)] = V2_mean
        print('WriteDataThisRow(): cell H{}: {}'.format(str(row), V2_mean), file=self.log)

        V2_sd = np.std(self.V2Data, ddof=1)[0]
        self.ws['I'+str(row)] = V2_sd
        print('WriteDataThisRow(): cell I{}: {}'.format(str(row), V2_sd), file=self.log)

        av_t = np.mean(self.VdTimes)[0]
        t_stamp = str(dt.datetime.fromtimestamp(av_t).strftime(fmt))
        self.ws['M'+str(row)] = t_stamp
        print('WriteDataThisRow(): cell M{}: {}'.format(str(row), t_stamp), file=self.log)

        Vd_mean = np.mean(self.VdData)[0]
        self.ws['N'+str(row)] = Vd_mean
        print('WriteDataThisRow(): cell N{}: {}'.format(str(row), Vd_mean), file=self.log)

        Vd_sd = np.std(self.VdData, ddof=1)[0]
        self.ws['O'+str(row)] = Vd_sd
        print('WriteDataThisRow(): cell O{}: {}'.format(str(row), Vd_sd), file=self.log)

        if devices.ROLES_INSTR['DVMT1'].demo is True:
            T1dvm_out = np.random.normal(108.0, 1.0e-2)
            self.ws['S' + str(row)].font = Font(color=colors.RED)
            # self.ws['S' + str(row)] = T1dvm_out
            # print('WriteDataThisRow(): cell S{}: {}'.format(str(row), T1dvm_out), file=self.log)
        else:
            T1dvm_out = devices.ROLES_INSTR['DVMT1'].send_cmd('READ?')
            self.ws['S' + str(row)].font = Font(color=colors.BLACK)
        self.ws['S'+str(row)] = T1dvm_out  # float(filter(self.filt, T1dvm_out))
        print('WriteDataThisRow(): cell S{}: {}'.format(str(row), T1dvm_out), file=self.log)

        if devices.ROLES_INSTR['DVMT2'].demo is True:
            T2dvm_out = np.random.normal(108.0, 1.0e-2)
            self.ws['T' + str(row)].font = Font(color=colors.RED)
            # self.ws['T' + str(row)] = T2dvm_out
            # print('WriteDataThisRow(): cell T{}: {}'.format(str(row), T2dvm_out))
        else:
            T2dvm_out = devices.ROLES_INSTR['DVMT2'].send_cmd('READ?')
            self.ws['T' + str(row)].font = Font(color=colors.BLACK)
        self.ws['T'+str(row)] = T2dvm_out  # float(filter(self.filt, T2dvm_out))
        print('WriteDataThisRow(): cell T{}: {}'.format(str(row), T2dvm_out), file=self.log)

        self.ws['U'+str(row)] = self.T1
        print('WriteDataThisRow(): cell U{}: {}'.format(str(row), self.T1), file=self.log)
        self.ws['V'+str(row)] = self.T2
        print('WriteDataThisRow(): cell V{}: {}'.format(str(row), self.T2), file=self.log)
        self.ws['W'+str(row)] = self.Troom
        print('WriteDataThisRow(): cell W{}: {}'.format(str(row), self.Troom), file=self.log)
        self.ws['X'+str(row)] = self.Proom
        print('WriteDataThisRow(): cell X{}: {}'.format(str(row), self.Proom), file=self.log)
        self.ws['Y'+str(row)] = self.RHroom
        print('WriteDataThisRow(): cell Y{}: {}'.format(str(row), self.RHroom), file=self.log)
        self.ws['Z'+str(row)] = self.Comment
        print('WriteDataThisRow(): cell Z{}: {}'.format(str(row), self.Comment), file=self.log)

        # Save after every row
        self.wb_io.save(self.xlfilename)

    def abort_run(self):
        # prematurely end run, prompted by regular checks of _want_abort flag
        self.standby()  # Set sources to 0V and leave system safe
        stop_ev = evts.DataEvent(t='-', Vm='-', Vsd='-', P=0, r='-', flag='E')  # End
        wx.PostEvent(self.RunPage, stop_ev)
        self.reset_buttons()

    def finish_run(self):
        # Run complete - leave system safe and final xl save
        self.wb_io.save(self.xlfilename)
        self.standby()  # Set sources to 0V and leave system safe
        stop_ev = evts.DataEvent(t='-', Vm='-', Vsd='-', P=0, r='-', flag='F')  # Finished
        wx.PostEvent(self.RunPage, stop_ev)
        stat_ev = evts.StatusEvent(msg='RUN COMPLETED', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)
        self.reset_buttons()

    def reset_buttons(self):
        self.RunPage.StartBtn.Enable(True)
        self.RunPage.RLinkBtn.Enable(True)
        self.RunPage.StopBtn.Enable(False)

    def standby(self):
        # Set sources to 0V and disable outputs
        devices.ROLES_INSTR['SRC1'].send_cmd('R0=')
        self.RunPage.V1Setting.SetValue(str(0))
        self.RunPage.V2Setting.SetValue(str(0))

    def abort(self):
        """abort worker thread."""
        # Method for use by main thread to signal an abort
        stat_ev = evts.StatusEvent(msg='abort(): Run aborted', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        self._want_abort = 1

    def filt(self, char):
        # A helper function to filter rubbish from DVM o/p unicode string
        # ...and retain any number (as a string)
        accept_str = u'-0.12345678eE9'
        return char in accept_str  # Returns 'True' or 'False'

"""--------------End of Thread class definition-------------------"""

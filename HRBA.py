# -*- coding: utf-8 -*-
"""
HRBA.py High Resistance Bridge Analysis (using imported GTC functions).

This script analyses data collected in an Excel file, generated by
the High Resistance Bridge Control (HRBC) application. Each 4-line
block of data in the 'Data' worksheet represents one measurement of
the unknown resistor R2('Rx') under a unique set of conditions
(time t, Temperature T, Voltage V).

Each HRBA run analyses one HRBC run, which normally consists of several
measurements at alternating low or high voltage (LV,HV). HRBA searches
the 'Rlink' worksheet for a block of Rlink data that has a matching run
ID and uses this information to calculate the link resistance Rd.

Where two temperature measurements of a resistor have been recorded
(eg with a GMH probe and a DVM monitoring a Pt sensor), the difference
is used to define a zero-valued type-B uncertainty that describes the
temperature definition. Where only a GMH reading is available the
temperature definition uncertainty defaults to 0 +/- 0.01 C with 4
degrees of freedom.

The multiple results from one analysis run are combined using least
squares fits to time, Temperature and Voltage, yielding six overall
measurements of R2 (at mean t, mean T, mean V for LV or HV conditions).

The fitting procedure also generates temperature and voltage coefficients
of resistance for the unknown resistor and an estimate of its drift rate.
NOTE: No correlations between these quantities are assumed.
All results are written to the 'Results' worksheet.

Created on Fri Sep 18 14:01:18 2015

@author: t.lawson
"""

import os
import sys
# sys.path.append("C:\Python27\Lib\site-packages\GTC")

import datetime as dt
import math

from openpyxl import load_workbook, utils  # cell
# from openpyxl.cell import get_column_letter #column_index_from_string
# from openpyxl.utils import get_column_letter

import GTC

import R_info  # useful functions

VERSION = 1.3

# DVM, GMH Correction factors, etc.

ZERO = GTC.ureal(0, 0)
FRAC_TOLERANCE = {'R2': 2e-2, 'G': 0.01, 'R1': 1}  # {'R2': 2e-4, 'G': 0.01, 'R1': 5e-2}
PPM_TOLERANCE = {'R2': 2e-4, 'G': 0.01, 'R1': 1e-3}
RLINK_MAX = 2000  # Ohms

datadir = input('Path to data directory:')
xlname = input('Excel filename:')
xlfile = os.path.join(datadir, xlname)

logname = R_info.Make_Log_Name(VERSION)
logfile = os.path.join(datadir, logname)
log = open(logfile, 'a')

now_tup = dt.datetime.now()
now_fmt = now_tup.strftime('%d/%m/%Y %H:%M:%S')
log.write(now_fmt + '\n' + xlfile + '\n')

# open existing workbook
print(str(xlfile))
wb_io = load_workbook(xlfile, data_only=True)  # data-only option added for validation
ws_Data = wb_io['Data']  # wb_io.get_sheet_by_name('Data')
ws_Rlink = wb_io['Rlink']  # wb_io.get_sheet_by_name('Rlink')
ws_Summary = wb_io['Results']  # wb_io.get_sheet_by_name('Results')
ws_Params = wb_io['Parameters']  # wb_io.get_sheet_by_name('Parameters')

# Get local parameters
Data_start_row = ws_Data.cell(row=1, column=2).value  # ['B1'].value
Data_stop_row = ws_Data.cell(row=2, column=2).value  # ['B2'].value
assert Data_start_row <= Data_stop_row, 'Stop row must follow start row!'

# Get instrument assignments
N_ROLES = 10  # 10 roles in total
role_descr = {}
for row in range(Data_start_row, Data_start_row + N_ROLES):
    # Read {role:description}
    # temp_dict = {ws_Data['AC' + str(row)].value: ws_Data['AD' + str(row)].value}
    # assert temp_dict.keys()[-1] is not None, 'Instrument assignment: Missing role!'
    # assert temp_dict.values()[-1] is not None, 'Instrument assignment: Missing description!'
    key = ws_Data.cell(row=row, column=29).value  # ['AC' + str(row)].value
    val = ws_Data.cell(row=row, column=30).value  # ['AD' + str(row)].value
    assert key is not None, 'Instrument assignment: Missing role!'
    assert val is not None, 'Instrument assignment: Missing description!'
    temp_dict = {key: val}
    role_descr.update(temp_dict)
    if ws_Data.cell(row=row, column=29).value == u'DVM12':  # ['AC'+str(row)].value
        range_mode = ws_Data.cell(row=row, column=31).value  # ['AE'+str(row)].value
        print('Range mode:', range_mode)

# ------------------------------------------------------------------- #
# _____________Extract resistor and instrument parameters____________ #

print('Reading parameters...')
log.write('Reading parameters...')
headings = (u'Resistor Info:', u'Instrument Info:',
            u'description', u'parameter', u'value',
            u'uncert', u'dof', u'label', u'Comment / Reference')

# Determine colummn indices from column letters:
col_A = utils.cell.column_index_from_string('A') - 1
col_B = utils.cell.column_index_from_string('B') - 1
col_C = utils.cell.column_index_from_string('C') - 1
col_D = utils.cell.column_index_from_string('D') - 1
col_E = utils.cell.column_index_from_string('E') - 1
col_F = utils.cell.column_index_from_string('F') - 1
col_G = utils.cell.column_index_from_string('G') - 1

col_I = utils.cell.column_index_from_string('I') - 1
col_J = utils.cell.column_index_from_string('J') - 1
col_K = utils.cell.column_index_from_string('K') - 1
col_L = utils.cell.column_index_from_string('L') - 1
col_M = utils.cell.column_index_from_string('M') - 1
col_N = utils.cell.column_index_from_string('N') - 1
col_O = utils.cell.column_index_from_string('O') - 1

R_params = []
R_row_items = []
I_params = []
I_row_items = []
R_values = []
I_values = []
R_DESCR = []
I_DESCR = []
R_sublist = []
I_sublist = []
last_I_row = last_R_row = 0

for r in ws_Params.rows:  # a tuple of row objects
    R_end = 0

    # description, parameter, value, uncert, dof, label:
    R_row_items = [r[col_A].value, r[col_B].value, r[col_C].value,
                   r[col_D].value, r[col_E].value, r[col_F].value,
                   r[col_G].value]

    I_row_items = [r[col_I].value, r[col_J].value, r[col_K].value,
                   r[col_L].value, r[col_M].value, r[col_N].value,
                   r[col_O].value]

    if R_row_items[0] is None:  # end of R_list
        R_end = 1

    # check this row for heading text
    if any(i in I_row_items for i in headings):
        continue  # Skip headings

    else:  # not header - main data
        # Get instrument parameters first...
        last_I_row = r[col_I].row
        I_params.append(I_row_items[1])
        I_values.append(R_info.Uncertainize(I_row_items))
        if I_row_items[1] == u'test':  # last parameter for this description
            I_DESCR.append(I_row_items[0])  # build description list
            I_sublist.append(dict(zip(I_params, I_values)))  # add parameter dictionary to sublist
            del I_params[:]
            del I_values[:]

        # Now attend to resistor parameters...
        if R_end == 0:  # Check we're not at the end of resistor data-block
            last_R_row = r[col_A].row  # Need to know this if we write more data, post-analysis
            R_params.append(R_row_items[1])
            R_values.append(R_info.Uncertainize(R_row_items))
            if R_row_items[1] == u'T_sensor':  # last parameter for this description
                R_DESCR.append(R_row_items[0])  # build description list
                R_sublist.append(dict(zip(R_params, R_values)))  # add parameter dictionary to sublist
                del R_params[:]
                del R_values[:] 

# Compile into dictionaries
"""
There are two dictionaries; one for instruments (I_INFO) and one for resistors
(R_INFO). Each dictionary item is keyed by the description (name) of the
instrument (resistor). Each dictionary value is itself a dictionary, keyed by
parameter, such as 'address' (for an instrument) or 'R_LV' (for a resistor
value, measured at 'low voltage').

"""
I_INFO = dict(zip(I_DESCR, I_sublist))
print(f'Found {len(I_INFO)} instruments ({last_I_row} rows)')
log.write(f'\nFound {len(I_INFO)} instruments ({last_I_row} rows)')

R_INFO = dict(zip(R_DESCR, R_sublist))
print(f'Found {len(R_INFO)} resistors ({last_R_row} rows)')
log.write(f'\nFound {len(R_INFO)} resistors ({last_R_row} rows)')

# -------------End of parameter extraction-------------- #
# ###################################################### #


# Determine the meanings of 'LV' and 'HV'
V1set_a = abs(ws_Data.cell(row=Data_start_row, column=1).value)  # ['A'+str(Data_start_row)].value
assert V1set_a is not None, 'Missing initial V1 value!'
V1set_b = abs(ws_Data.cell(row=Data_start_row+4, column=1).value)  # ['A'+str(Data_start_row+4)].value
assert V1set_b is not None, 'Missing second V1 value!'

if V1set_a < V1set_b:
    LV = V1set_a
    HV = V1set_b
elif V1set_b < V1set_a:
    LV = V1set_b
    HV = V1set_a
else:  # 'HV' and 'LV' equal
    LV = HV = V1set_a
print(f'LV ={LV}; HV ={HV}')

# Set up reading of Data sheet
Data_row = Data_start_row

# Get start_row on Summary sheet
summary_start_row = ws_Summary.cell(row=1, column=2).value
assert summary_start_row is not None, 'Missing start row on Results sheet!'

# Get run identifier and copy to Results sheet
Run_Id = ws_Data.cell(row=Data_start_row-1, column=2).value
assert Run_Id is not None, 'Missing Run Id!'

ws_Summary['C'+str(summary_start_row)] = 'Run Id:'
ws_Summary['D'+str(summary_start_row)] = str(Run_Id)

# Get run comment and extract R names & R values
Data_comment = ws_Data.cell(row=Data_row, column=26).value
assert Data_comment is not None, 'Missing Comment!'

print(Data_comment)
log.write(f'\n{Data_comment}')
print('Run Id:', Run_Id)
log.write(f'\nRun Id: {Run_Id}')

# Write headings
summary_row = R_info.WriteHeadings(ws_Summary, summary_start_row, VERSION)

# Lists of dictionaries...
# ...(with name,time,Resistance,Temperature,Voltage entries).
results_HV = []  # High voltage measurements
results_LV = []  # Low voltage measurements

# Get resistor names and values
R1_name, R2_name = R_info.ExtractNames(Data_comment)
R1val = R_info.GetRval(R1_name)
R2val = R_info.GetRval(R2_name)

# Check for knowledge of R2:
if R2_name not in R_INFO:
    sys.exit('ERROR - Unknown Rs: '+R2_name)


# ## __________Get Rd value__________## #
# 1st, detetermine data format
N_revs = ws_Rlink.cell(row=2, column=2).value  # Number of reversals = number of columns
assert N_revs is not None and N_revs > 0, 'Missing or no reversals!'
N_reads = ws_Rlink.cell(row=3, column=2).value  # Number of readings = number of rows
assert N_reads is not None and N_reads > 0, 'Missing or no reads!'
head_height = 6  # Rows of header before each block of data
jump = head_height + N_reads  # rows to jump between starts of each header

# Find correct RLink data-header
RL_start_row = R_info.GetRLstartrow(ws_Rlink, Run_Id, jump, log)
assert RL_start_row > 1, 'Unable to find matching Rlink data!'

# Next, define nom_R,abs_V quantities
"""
Assume all 'nominal' values have 100 ppm std.uncert. with 8 dof.
"""
val1 = ws_Rlink.cell(row=RL_start_row+2, column=3).value
assert val1 is not None, 'Missing nominal R1 value!'
nom_R1 = GTC.ureal(val1, val1/1e4, 8, label='nom_R1')  # don't know uncertainty of nominal values
val2 = ws_Rlink.cell(row=RL_start_row+3, column=3).value
assert val2 is not None, 'Missing nominal R2 value!'
nom_R2 = GTC.ureal(val2, val2/1e4, 8, label='nom_R2')  # don't know uncertainty of nominal values
val1 = ws_Rlink.cell(row=RL_start_row+2, column=4).value
assert val1 is not None, 'Missing nominal V1 value!'
abs_V1 = GTC.ureal(val1, val1/1e4, 8, label='abs_V1')  # don't know uncertainty of nominal values
val2 = ws_Rlink.cell(row=RL_start_row+3, column=4).value
assert val2 is not None, 'Missing nominal V2 value!'
abs_V2 = GTC.ureal(val2, val2/1e4, 8, label='abs_V2')  # don't know uncertainty of nominal values

# Calculate I
I = GTC.result((abs_V1 + abs_V2) / (nom_R1 + nom_R2), 'Rd_I' + Run_Id)
# I = (abs_V1 + abs_V2) / (nom_R1 + nom_R2)
# I.label = 'Rd_I' + Run_Id

# Average all +Vs and -Vs
Vp = []  # Positive polarity measurements
Vn = []  # Negative polarity measurements

for Vrow in range(RL_start_row+5, RL_start_row+5+N_reads):

    col = 1
    while col <= N_revs:  # cycle through cols 1 to N_revs
        Vp.append(ws_Rlink[utils.get_column_letter(col)+str(Vrow)].value)
        assert Vp[-1] is not None, 'Missing Vp value!'
        col += 1

        Vn.append(ws_Rlink[utils.get_column_letter(col)+str(Vrow)].value)
        assert Vn[-1] is not None, 'Missing Vn value!'
        col += 1

av_dV_p = GTC.result(GTC.ta.estimate(Vp), 'av_dV_p' + Run_Id)
# av_dV_p = GTC.ta.estimate(Vp)
# av_dV_p.label = 'av_dV_p' + Run_Id
av_dV_n = GTC.result(GTC.ta.estimate(Vn), 'av_dV_n' + Run_Id)
# av_dV_n = GTC.ta.estimate(Vn)
# av_dV_n.label = 'av_dV_n' + Run_Id
av_dV = GTC.result(0.5*GTC.magnitude(av_dV_p - av_dV_n), 'Rd_dV' + Run_Id)
# av_dV = 0.5*GTC.magnitude(av_dV_p - av_dV_n)
# av_dV.label = 'Rd_dV' + Run_Id

# Finally, calculate Rd
Rd = GTC.result(av_dV/I, label='Rlink ' + Run_Id)
print(f'\nRlink = {Rd.x} +/- {Rd.u}, dof = {Rd.df}')
assert Rd.x < RLINK_MAX, f'High link resistance! {Rd.x} Ohm'
# assert Rd.x > Rd.u, 'Link resistance uncertainty > value!'
log.write(f'\nRlink = {Rd.x} +/- {Rd.u}, dof = {Rd.df}')

# ##__________End of Rd section___________## #

raw_gmh1 = []
raw_gmh2 = []
T_dvm1 = []
T_dvm2 = []
R_dvm1 = []
R_dvm2 = []
times = []
RHs = []
Ps = []
Ts = []

##############################
# ___Loop over data rows ___ #
print('\nLooping over data rows', Data_start_row, 'to', Data_stop_row, '...')
log.write('\nLooping over data rows ' + str(Data_start_row) + ' to ' + str(Data_stop_row) + '\n')
while Data_row <= Data_stop_row:
    # R2 parameters:
    V2set = ws_Data.cell(row=Data_row, column=2).value  # Changed from Data_start_row!
    assert V2set is not None, 'Missing V2 setting!'
    V1set = ws_Data.cell(row=Data_row, column=1).value  # Changed from Data_start_row!
    assert V1set is not None, 'Missing V1 setting!'

    # Select R2 info based on applied voltage ('LV' or 'HV')
    Vdif_LV = abs(abs(V2set)-R_INFO[R2_name]['VRef_LV'])
    Vdif_HV = abs(abs(V2set)-R_INFO[R2_name]['VRef_HV'])
    if Vdif_LV < Vdif_HV:
        R2_0 = R_INFO[R2_name]['R0_LV']
        R2TRef = R_INFO[R2_name]['TRef_LV']
        R2VRef = R_INFO[R2_name]['VRef_LV']
    elif Vdif_LV > Vdif_HV:
        R2_0 = R_INFO[R2_name]['R0_HV']
        R2TRef = R_INFO[R2_name]['TRef_HV']
        R2VRef = R_INFO[R2_name]['VRef_HV']
    else:
        R2_0 = R_INFO[R2_name]['R0_LV']
        R2TRef = R_INFO[R2_name]['TRef_LV']
        R2VRef = R_INFO[R2_name]['VRef_LV']

    # Select appropriate value of VRC, etc.
    """
    #################################################################
    NOTE: Now replace VRCs with individual gain factors for
    each test-V (at mid- or top-of-range), on each instrument. Since
    this matches available info in DMM cal. cert. and minimises the
    number of possible values (ie: No. of test-Vs] < [No. of possible
    voltage ratios]).
    #################################################################
    """
#    G1_code = R_info.Vgain_codes_auto[round(V1set,1)]
#    G1 = I_INFO[role_descr['DVM12']][G1_code]

    V2rnd = math.pow(10, round(math.log10(abs(V2set))))  # Rnd to nearest 10-pwr
    V1rnd = math.pow(10, round(math.log10(abs(V1set))))
    if 'AUTO' in range_mode:
        G2_code = R_info.Vgain_codes_auto[V2rnd]
        G1_code = R_info.Vgain_codes_auto[V1rnd]
    else:  # 'FIXED'
        if round(V1set) >= round(abs(V2set)):
            G1_code = R_info.Vgain_codes_auto[V1rnd]
            G2_code = R_info.Vgain_codes_fixed[V2rnd]
        else:
            G2_code = R_info.Vgain_codes_auto[V2rnd]
            G1_code = R_info.Vgain_codes_fixed[V1rnd]

    G1 = I_INFO[role_descr['DVM12']][G1_code]
    G2 = I_INFO[role_descr['DVM12']][G2_code]

    vrc = GTC.result(G2/G1, label='vrc ' + Run_Id)

    Vlin_pert = I_INFO[role_descr['DVMd']]['linearity_pert']  # linearity used in G calculation
    Vlin_Vdav = I_INFO[role_descr['DVMd']]['linearity_Vdav']  # linearity used in Vd calculation

    # Start list of influence variables
    influencies = [Rd, G1, G2, Vlin_pert,
                   Vlin_Vdav, R2TRef, R2VRef]  # R2 dependencies

    R2alpha = R_INFO[R2_name]['alpha']
    R2beta = R_INFO[R2_name]['beta']
    R2gamma = R_INFO[R2_name]['gamma']
    R2Tsensor = R_INFO[R2_name]['T_sensor']
    influencies.extend([R2_0, R2alpha, R2beta, R2gamma])  # R2 dependencies

    if R1_name not in R_INFO:
        R1Tsensor = 'Pt 100r'  # assume a Pt sensor in unknown resistor
    else:
        R1Tsensor = R_INFO[R1_name]['T_sensor']

    # GMH correction factors
    GMH1_cor = I_INFO[role_descr['GMH1']]['T_correction']
    GMH2_cor = I_INFO[role_descr['GMH2']]['T_correction']

    # Temperature measurement, RH and times:
    del raw_gmh1[:]  # list for 4 corrected T1 gmh readings
    del raw_gmh2[:]  # list for 4 corrected T2 gmh readings
    del T_dvm1[:]  # list for 4 corrected T1(dvm) readings
    del T_dvm2[:]  # list for 4 corrected T2(dvm) readings
    del R_dvm1[:]  # list for 4 corrected dvm readings
    del R_dvm2[:]  # list for 4 corrected dvm readings
    del times[:]  # list for 3*4 mean measurement time-strings
    del RHs[:]  # list for 4 RH values
    del Ps[:]  # list for 4 room pressure values
    del Ts[:]  # list for 4 room Temp values

    # Process times, RH and temperature data in this 4-row block:
    for r in range(Data_row, Data_row+4):  # build list of 4 gmh / T-probe dvm readings
        assert ws_Data.cell(row=r, column=21).value is not None, 'No R1 GMH temperature data!'
        assert ws_Data.cell(row=r, column=22)['V'+str(r)].value is not None, 'No R2 GMH temperature data!'
        raw_gmh1.append(ws_Data.cell(row=r, column=21).value)
        raw_gmh2.append(ws_Data.cell(row=r, column=22).value)

        assert ws_Data.cell(row=r, column=7).value is not None, 'No V2 timestamp!'
        assert ws_Data.cell(row=r, column=13)['M'+str(r)].value is not None, 'No Vd1 timestamp!'
        assert ws_Data.cell(row=r, column=16)['P'+str(r)].value is not None, 'No V1 timestamp!'
        times.append(ws_Data.cell(row=r, column=7).value)
        times.append(ws_Data.cell(row=r, column=13).value)
        times.append(ws_Data.cell(row=r, column=16).value)

        assert ws_Data.cell(row=r, column=19).value is not None, 'No R1 raw DVM (temperature) data!'
        raw_dvm1 = ws_Data.cell(row=r, column=19).value

        assert ws_Data.cell(row=r, column=20).value is not None, 'No R2 raw DVM (temperature) data!'
        raw_dvm2 = ws_Data.cell(row=r, column=20).value

        # Check corrections for range-dependant values...
        # and apply appropriate corrections
        assert raw_dvm1 > 0, 'DVMT1: Negative resistance value!'
        assert raw_dvm2 > 0, 'DVMT2: Negative resistance value!'
        if raw_dvm1 < 120:
            T1DVM_cor = I_INFO[role_descr['DVMT1']]['correction_100r']
        elif raw_dvm1 < 12e3:
            T1DVM_cor = I_INFO[role_descr['DVMT1']]['correction_10k']
        else:
            T1DVM_cor = I_INFO[role_descr['DVMT1']]['correction_100k']
        R_dvm1.append(raw_dvm1*(1+T1DVM_cor))

        if raw_dvm2 < 120:
            T2DVM_cor = I_INFO[role_descr['DVMT2']]['correction_100r']
        elif raw_dvm2 < 12e3:
            T2DVM_cor = I_INFO[role_descr['DVMT2']]['correction_10k']
        else:
            T2DVM_cor = I_INFO[role_descr['DVMT2']]['correction_100k']
        R_dvm2.append(raw_dvm2*(1+T2DVM_cor))

    # Mean temperature from GMH
    # Data are plain numbers (with digitization rounding),
    # so use ta.estimate_digitized() to return a ureal.
    assert len(raw_gmh1) > 1, 'Not enough GMH1 temperatures to average!'
    T1_av_gmh = GTC.result(GTC.ta.estimate_digitized(raw_gmh1, 0.01) + GMH1_cor,
                           label='T1_av_gmh ' + Run_Id)

    assert len(raw_gmh2) > 1, 'Not enough GMH2 temperatures to average!'
    T2_av_gmh = GTC.result(GTC.ta.estimate_digitized(raw_gmh2, 0.01) + GMH2_cor,
                           label='T2_av_gmh ' + Run_Id)

    assert len(times) > 1, 'Not enough timestamps to average!'
    times_av_str = R_info.av_t_strin(times, 'str')  # mean time(as a time string)
    times_av_fl = R_info.av_t_strin(times, 'fl')  # mean time(as a float)

    """
    TO DO: Incorporate ambient T, P, %RH readings into final reported results...
    
    assert len(RHs) > 1,'Not enough RH values to average!'
    # Digitization could be 2 or 3 decimal places, depending on RH probe:
    RH_av = GTC.ar.result(GTC.ta.estimate_digitized(RHs,R_info.GetDigi(RHs)),label = 'RH_av')
    
    ... (and same for T, P) ...

    """

    # Build lists of 4 temperatures (calculated from T-probe dvm readings)...
    # ... and calculate mean temperatures
    if R1Tsensor in ('none', 'any'):  # no or unknown T-sensor (Tinsleys or T-sensor itelf)
        T_dvm1 = [ZERO, ZERO, ZERO, ZERO]
    else:
        assert len(R_dvm1) > 1, 'Not enough R_dvm1 values to average!'
        for R in R_dvm1:  # convert resistance measurement to a temperature
            T_dvm1.append(R_info.R_to_T(R_INFO[R1Tsensor]['alpha'],
                                        R_INFO[R1Tsensor]['beta'], R,
                                        R_INFO[R1Tsensor]['R0_LV'],
                                        R_INFO[R1Tsensor]['TRef_LV']))
    if R2Tsensor in ('none', 'any'):
        T_dvm2 = [ZERO, ZERO, ZERO, ZERO]
    else:
        assert len(R_dvm2) > 1, 'Not enough R_dvm2 values to average!'
        for R in R_dvm2:  # convert resistance measurement to a temperature
            T_dvm2.append(R_info.R_to_T(R_INFO[R2Tsensor]['alpha'],
                                        R_INFO[R2Tsensor]['beta'], R,
                                        R_INFO[R2Tsensor]['R0_LV'],
                                        R_INFO[R2Tsensor]['TRef_LV']))

    # Mean temperature from T-probe dvm
    # Data are high-precision plain numbers,
    # so use ta.estimate() to return a ureal.
    # T1_av_dvm = GTC.result(GTC.ta.estimate(T_dvm1),
    #                           label='T1_av_dvm' + Run_Id)
    # T2_av_dvm = GTC.result(GTC.ta.estimate(T_dvm2),
    #                           label='T2_av_dvm' + Run_Id)

    # Mean temperatures and temperature definitions
#    if role_descr['DVMT1']=='none':  # No aux. T sensor or DVM not associated with R1 (just GMH)
    T1_av = T1_av_gmh
    T1_av_dvm = GTC.ureal(0, 0)  # ignore any dvm data
    Diff_T1 = GTC.ureal(0, 0)  # No temperature disparity (GMH only)
#    else:
#        T1_av = GTC.ar.result(GTC.fn.mean((T1_av_dvm,T1_av_gmh)),label='T1_av'+ Run_Id)
#        Diff_T1 = GTC.magnitude(T1_av_dvm-T1_av_gmh)

#    if role_descr['DVMT2']=='none':  # No aux. T sensor or DVM not associated with R2 (just GMH)
    T2_av = T2_av_gmh
    T2_av_dvm = GTC.ureal(0, 0)  # ignore any dvm data
    Diff_T2 = GTC.ureal(0, 0)  # No temperature disparity (GMH only)
    influencies.append(T2_av_gmh)  # R2 dependancy
#    else:
#        T2_av = GTC.ar.result( GTC.fn.mean((T2_av_dvm,T2_av_gmh)),label='T2_av' + Run_Id)
#        Diff_T2 = GTC.ar.result(GTC.magnitude(T2_av_dvm-T2_av_gmh),label='Diff_T2' + Run_Id)
#        influencies.append(T2_av_dvm,T2_av_gmh) # R2 dependancy

    # Default T definition arises from imperfect positioning of sensors wrt resistor:
    T_def = GTC.ureal(0, GTC.type_b.distribution['gaussian'](0.01), 3,
                      label='T_def ' + Run_Id)

    # T-definition arises from imperfect positioning of both probes AND their disagreement:
    T_def1 = GTC.result(GTC.ureal(0, Diff_T1.u/2, 7) + T_def,
                        label='T_def1 ' + Run_Id)
    T_def2 = GTC.result(GTC.ureal(0, Diff_T2.u/2, 7) + T_def,
                        label='T_def2 ' + Run_Id)
    influencies.append(T_def2)  # R2 dependancy

    # Raw voltage measurements: V: [Vp,Vm,Vpp,Vppp]
    # All readings are precise enough not to worry about digitization error...
    V1 = []
    V2 = []
    Vd = []
    for line in range(4):

        V1.append(GTC.ureal(ws_Data.cell(row=Data_row+line, column=17).value,
                            ws_Data.cell(row=Data_row+line, column=18).value,
                            ws_Data.cell(row=Data_row+line, column=3).value-1,
                            label='V1_'+str(line) + ' ' + Run_Id))
        V2.append(GTC.ureal(ws_Data.cell(row=Data_row+line, column=8).value,
                            ws_Data.cell(row=Data_row+line, column=9).value,
                            ws_Data.cell(row=Data_row+line, column=3).value-1,
                            label='V2_'+str(line) + ' ' + Run_Id))
        Vd.append(GTC.ureal(ws_Data.cell(row=Data_row+line, column=14).value,
                            ws_Data.cell(row=Data_row+line, column=15).value,
                            ws_Data.cell(row=Data_row+line, column=3).value-1,
                            label='Vd_'+str(line) + ' ' + Run_Id))
        assert V1[-1] is not None, 'Missing V1 data!'
        assert V2[-1] is not None, 'Missing V2 data!'
        assert Vd[-1] is not None, 'Missing Vd data!'
    influencies.extend(V1+V2+Vd)  # R2 dependancies - raw measurements

    # Define drift
    uncert = abs(Vd[2]-(Vd[0]+((Vd[3]-Vd[2])/(V2[3]-V2[2]))*(V2[2]-V2[0])))/4
    Vdrift1 = GTC.ureal(0, GTC.tb.distribution['gaussian'](uncert), 8,
                        label='Vdrift_pert ' + Run_Id)
    Vdrift2 = GTC.ureal(0, GTC.tb.distribution['gaussian'](uncert), 8,
                        label='Vdrift_Vdav ' + Run_Id)

    Vdrift = {'pert': Vdrift1, 'Vdav': Vdrift2}
    influencies.extend([Vdrift['pert'], Vdrift['Vdav']])  # R2 dependancies

    # Mean voltages
    V1av = (V1[0]-2*V1[1]+V1[2])/4
    V2av = (V2[0]-2*V2[1]+V2[2])/4
    Vdav = (Vd[0]-2*Vd[1]+Vd[2])/4 + Vlin_Vdav + Vdrift['Vdav']

    # Effect of v2 perturbation
    delta_Vd = Vd[3] - Vd[2] + Vlin_pert + Vdrift['pert']
    delta_V2 = V2[3] - V2[2]

    # Calculate R2 (corrected for T and V)
    dT2 = T2_av - R2TRef + T_def2

    # NOTE: NEED TWO abs() TO ENSURE NON-NEGATIVE DIFFERENCE:
    dV2 = abs(abs(V2av) - R2VRef)

    R2 = R2_0*(1 + R2alpha * dT2 + R2beta * dT2 ** 2 + R2gamma * dV2) + Rd
    print(f'R2 = {R2.x} +/- {R2.u}, dof = {R2.df}')
    # print 'R2 =',GTC.summary(R2)
    frac_err = abs(R2.x-R2val) / R2val
    assert frac_err < FRAC_TOLERANCE['R2'], f'R2 > 100 ppm from nominal! R2 ({frac_err})'
    # assert abs(R2.x-R2val)/R2val < PPM_TOLERANCE['R2'], 'R2 > 100 ppm from nominal! R2 = {0}'.format(R2.x)

    # calculate R1
    R1 = R2*vrc*V1av*delta_Vd/(Vdav*delta_V2 - V2av*delta_Vd)
    print(f'R1 = {R1.x} +/- {R1.u}, dof = {R1.df}')
    frac_err = abs(R1.x - R1val) / R1val
    assert frac_err < FRAC_TOLERANCE['R1'], f'R1 > 1000 ppm from nominal ({frac_err})!'
    # assert abs(R1.x-R1val)/R1val < PPM_TOLERANCE['R1'], 'R1 > 1000 ppm from nominal!'

    T1 = T1_av + T_def1
    print(f'{R1} at temperature {T1}')

    # Combine data for this measurement: name,time,R,T,V and write to Summary sheet:
    this_result = {'name': R1_name, 'time_str': times_av_str,
                   'time_fl': times_av_fl, 'V': V1av, 'R': R1, 'T': T1,
                   'R_expU': R1.u*GTC.rp.k_factor(R1.df)}  # 'quick=False' arg deprecated.

    R_info.WriteThisResult(ws_Summary, summary_row, this_result)

    # build uncertainty budget table
    budget_table = []
    for i in influencies:  # rp.u_component(R1_gmh,i) gives + or - values
        if i.u > 0:
            sensitivity = GTC.rp.sensitivity(R1, i)
        else:
            sensitivity = 0
        budget_table.append([i.label, i.x, i.u, i.df, sensitivity,
                             GTC.component(R1, i)])

    budget_table_sorted = sorted(budget_table, key=R_info.by_u_cont,
                                 reverse=True)

    # write budget to Summary sheet
    summary_row = R_info.WriteBudget(ws_Summary, summary_row,
                                     budget_table_sorted)
    summary_row += 1  # Add a blank line between each measurement for ease of reading

    # Separate results by voltage (V1av) if different
    if HV == LV:
        results_LV.append(this_result)
        results_HV.append(this_result)
    elif abs(V1av.x - LV) < 1:
        results_LV.append(this_result)
    else:
        results_HV.append(this_result)

    del influencies[:]
    Data_row += 4  # Move to next measurement

# ----- End of data-row loop ---- #
# ############################### #


# At this point the summary row has reached its maximum for this analysis run
# ...so make a note of it, for use as the next run's starting row:
ws_Summary['B1'] = summary_row + 1  # Add extra row between runs

# Go back to the top of summary block, ready for writing run results
summary_row = summary_start_row + 1

########################################################################

"""
In the next section values of R1 are derived from fits to Temperature.
The Temperature data are offset so the mean is at ~zero, then the fits
are used to calculate R1 at the mean Temperature. LV and HV values are
obtained separately. The mean time, Temperature and Voltage values are
also reported.
"""

# Weighted total least-squares fit (R1-T), LV
print('\nLV:')
log.write('\nLV:')
R1_LV, Ohm_per_C_LV, T_LV, V_LV, date = R_info.write_R1_T_fit(results_LV,
                                                              ws_Summary,
                                                              summary_row, log)
alpha_LV = Ohm_per_C_LV/R1_LV

summary_row += 1

# Weighted total least-squares fit (R1-T), HV
print('\nHV:')
log.write('\nHV:')
R1_HV, Ohm_per_C_HV, T_HV, V_HV, date = R_info.write_R1_T_fit(results_HV,
                                                              ws_Summary,
                                                              summary_row, log)
alpha_HV = Ohm_per_C_HV/R1_HV

alpha = GTC.fn.mean([alpha_LV, alpha_HV])
beta = GTC.ureal(0, 0)  # assume no beta

if HV == LV:  # Can't estimate gamma
    gamma = GTC.ureal(0, 0)
else:
    gamma = ((R1_HV-R1_LV)/(V_HV-V_LV))/R1_LV

summary_row += 2

ws_Summary['R'+str(summary_row)] = 'alpha (/C)'
ws_Summary['V'+str(summary_row)] = 'gamma (/V)'

summary_row += 1

ws_Summary['R'+str(summary_row)] = alpha.x
ws_Summary['S'+str(summary_row)] = alpha.u

if math.isinf(alpha.df) or math.isnan(alpha.df):
    ws_Summary['T'+str(summary_row)] = str(alpha.df)
else:
    ws_Summary['T'+str(summary_row)] = round(alpha.df)

ws_Summary['V'+str(summary_row)] = gamma.x
ws_Summary['W'+str(summary_row)] = gamma.u
if math.isinf(gamma.df) or math.isnan(alpha.df):
    ws_Summary['X'+str(summary_row)] = str(gamma.df)
else:
    ws_Summary['X'+str(summary_row)] = round(gamma.df)

#######################################################################

"""
Finally, if R1 is a resistor that is not included in the 'parameters'
sheet it should be added to the 'current knowledge'...
"""

params = ['R0_LV', 'TRef_LV', 'VRef_LV', 'R0_HV', 'TRef_HV', 'VRef_HV',
          'alpha', 'beta', 'gamma', 'date', 'T_sensor']

R_data = [R1_LV, T_LV, V_LV, R1_HV, T_HV, V_HV, alpha, beta, gamma, date,
          'none']

if R1_name not in R_INFO:
    print(f'Adding {R1_name} to resistor info...')
    last_R_row = R_info.update_R_Info(R1_name, params, R_data, ws_Params,
                                      last_R_row, Run_Id, VERSION)
else:
    print(f'\nAlready know about {R1_name}')

# Save workbook
wb_io.save(xlfile)
print('_____________HRBA DONE_______________')
log.write('\n_____________HRBA DONE_______________\n\n')
log.close()

# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 09:36:42 2015

DEVELOPMENT VERSION

@author: t.lawson

RLink.py:
Thread class that executes processing.
Contains definitions for usual __init__() and run() methods
 AND an abort() method. The Run() method forms the core of the
 procedure - any changes to the way the measurements are taken
 should be made here, and within included subroutines.
"""

from threading import Thread
import datetime as dt
import time
# import string

import numpy as np
import wx
from openpyxl.utils import get_column_letter  # column_index_from_string
from openpyxl.styles import Font, colors

import HighRes_events as evts
import devices
import R_info


class RLThread(Thread):
    """RLink Thread Class."""
    def __init__(self, parent):
        # This runs when an instance of the class is created
        Thread.__init__(self)
        self.RunPage = parent
        self.SetupPage = self.RunPage.GetParent().GetPage(0)
        self.PlotPage = self.RunPage.GetParent().GetPage(2)
        self.TopLevel = self.RunPage.GetTopLevelParent()
        self.comment = self.RunPage.Comment.GetValue()
        self._want_abort = 0
        self.RLink_data = []

        self.log = self.SetupPage.log

        print('\nRole -> Instrument:')
        print('------------------------------')
        # Print all instrument objects
        for r in devices.ROLES_WIDGETS.keys():
            d = devices.ROLES_WIDGETS[r]['icb'].GetValue()
            print('{} -> {}'.format(devices.INSTR_DATA[d]['role'], d))
            if r != devices.INSTR_DATA[d]['role']:
                devices.INSTR_DATA[d]['role'] = r
                print('Role data corrected to: {} -> {}'.format(r, d))

        # Get filename of Excel file
        self.xlfilename = self.SetupPage.XLFile.GetValue()

        # Find existing workbook
        self.wb_io = self.SetupPage.wb
        self.ws = self.wb_io.get_sheet_by_name('Rlink')

        # read start row & run parameters from Excel file
        self.start_row = self.ws['B1'].value  # 1st row of actual data (after 6 lines of header)
        self.headrow = self.start_row - 6
        self.N_reversals = self.ws['B2'].value
        self.N_readings = self.ws['B3'].value

        self.AbsV1 = self.ws['D1'].value
        self.AbsV2 = self.ws['D2'].value
        self.MaxV = max(self.AbsV1, self.AbsV2)

        self.settle_time = self.RunPage.SettleDel.GetValue()

        self.R1Name = self.SetupPage.R1Name.GetValue()
        self.R2Name = self.SetupPage.R2Name.GetValue()

        self.V1set = 0.0
        self.V2set = 0.0
        self.Vdiff = 0.0

        # Extract resistor nominal values from names
        # r1multiplier = self.Getmultiplier(self.R1Name)
        # r2multiplier = self.Getmultiplier(self.R2Name)
        # self.R1Val = r1multiplier*int(string.strip(string.split(self.R1Name)[-1], string.letters))
        # self.R2Val = r2multiplier*int(string.strip(string.split(self.R2Name)[-1], string.letters))
        self.R1Val = R_info.get_r_val(self.R1Name)
        self.R2Val = R_info.get_r_val(self.R2Name)
        self.start()  # Starts the thread running on creation

    def run(self):
        # Run Worker Thread. This is where all the important stuff goes.

        # Set button availibility
        self.RunPage.StopBtn.Enable(True)
        self.RunPage.StartBtn.Enable(False)
        self.RunPage.RLinkBtn.Enable(False)

        stat_ev = evts.StatusEvent(msg='RLThread.run():', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='Waiting to settle...', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

        time.sleep(self.settle_time)

        stat_ev = evts.StatusEvent(msg='', field='b')  # write to both fields
        wx.PostEvent(self.TopLevel, stat_ev)

        # Define headings
        headrows = range(self.headrow, self.start_row)
        row_content = [['Run Id:', str(self.RunPage.run_id)],
                       ['Comment', self.comment],
                       [str(dt.datetime.today().strftime("%d/%m/%Y %H:%M:%S")),
                       '', 'Nom. value', '|V|'],
                       ['R1', self.R1Name, self.R1Val, self.AbsV1],
                       ['R2', self.R2Name, self.R2Val, self.AbsV2]]
        cap_delta = u'\N{GREEK CAPITAL LETTER DELTA}'
        last_head_row = []
        for c in range(1, 6):
            last_head_row.append(cap_delta + 'V+')
            last_head_row.append(cap_delta + 'V-')
        row_content.append(last_head_row)

        headings = dict(zip(headrows, row_content))

        for r in headings.keys():
            for c in range(1, len(headings[r])+1):
                if r == self.headrow + 5:  # 'delta_V' row
                    if c % 2 == 0:  # even columns
                        col = colors.BLUE
                    else:  # odd columns
                        col = colors.RED
                    self.ws.cell(row=r, column=c).font = Font(color=col)
                if r == self.headrow:  # 1st row (Run Id)
                    self.ws.cell(row=r, column=c).font = Font(b=True)
                self.ws.cell(row=r, column=c).value = headings[r][c-1]

        revs = 1

        # Configuration and initialisation
        devices.ROLES_INSTR['switchbox'].send_cmd(devices.SWITCH_CONFIGS['V2'])
        self.SetupPage.Switchbox.SetValue('V2')  # update sw'box config icb
        devices.ROLES_INSTR['DVMd'].send_cmd('FUNC DCV,AUTO')
        dvm_op = devices.ROLES_INSTR['DVMd'].read()  # Set appropriate V-range
        devices.ROLES_INSTR['DVMd'].send_cmd('DCV,' + str(dvm_op))
        devices.ROLES_INSTR['DVMd'].send_cmd('LFREQ LINE')
        devices.ROLES_INSTR['SRC1'].send_cmd('R0=')
        time.sleep(3)

        self.V1set = self.AbsV1
        self.V2set = self.AbsV2*-1
        self.Vdiff = self.V1set - self.V2set

        while revs <= self.N_reversals:  # column index
            if self._want_abort:
                self.abort_run()
                return

            del self.RLink_data[:]

            # Apply source voltages
            # - Voltage displays control sources
            self.RunPage.V1Setting.SetValue(str(self.V1set))
            time.sleep(5)
            self.RunPage.V2Setting.SetValue(str(self.V2set))
            if self._want_abort:
                self.abort_run()
                return            
            time.sleep(60)
            row = 1

            # Only store 10 readings per line, and then clear
            col_letter = get_column_letter(revs)
            d = devices.ROLES_WIDGETS['DVMd']['icb'].GetValue()
            while row <= self.N_readings:  # row index
                if devices.INSTR_DATA[d]['demo']:
                    dvm_op = np.random.normal(self.Vdiff*1.0e-6,
                                              abs(self.Vdiff*1.0e-8))
                    # self.RLink_data.append(dvm_op)
                else:
                    devices.ROLES_INSTR['DVMd'].send_cmd('LFREQ LINE')
                    time.sleep(1)
                    devices.ROLES_INSTR['DVMd'].send_cmd('AZERO ONCE')
                    time.sleep(1)  # was 10
                    dvm_op = devices.ROLES_INSTR['DVMd'].read()
                    # self.RLink_data.append(float(filter(self.filt, dvm_op)))
                self.RLink_data.append(dvm_op)

                prog = 100 * ((revs - 1) * self.N_readings + row) / (self.N_reversals * self.N_readings)  # % progress
                update_ev = evts.DataEvent(t=0, Vm=self.RLink_data[row-1],
                                           Vsd=0, P=prog, r=col_letter + str(row),
                                           flag='-')
                wx.PostEvent(self.RunPage, update_ev)
                if revs % 2 == 0:  # even columns
                    col = colors.BLUE
                else:  # odd columns
                    col = colors.RED
                self.ws.cell(row=self.start_row+row-1,
                             column=revs).font = Font(color=col)
                self.ws.cell(row=self.start_row+row-1,
                             column=revs).value = self.RLink_data[row-1]
                row += 1
            # (end of readings loop)

            print(self.RLink_data[row-2])

            # Reverse source polarities
            self.V1set *= -1
            self.V2set *= -1
            self.Vdiff = self.V1set - self.V2set

            revs += 1

            # Set next data-block's start row (allowing for gap+6-line header)
            self.ws['B1'] = self.start_row + self.N_readings + 7
        # (end of reversals loop)
        self.finish_run()
        return

    def abort_run(self):
        # prematurely end run
        self.standby()  # Set sources to 0V and leave system safe

        stat_ev = evts.StatusEvent(msg='AbortRun(): Run stopped', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)

        stop_ev = evts.DataEvent(t='-', Vm='-', Vsd='-', P=0, r='-',
                                 flag='E')  # End
        wx.PostEvent(self.RunPage, stop_ev)

        self.RunPage.RLinkBtn.Enable(True)
        self.RunPage.StartBtn.Enable(True)
        self.RunPage.StopBtn.Enable(False)

    def finish_run(self):
        # save data in xl file
        self.wb_io.save(self.xlfilename)

        self.standby()  # Set sources to 0V and leave system safe

        stop_ev = evts.DataEvent(t='-', Vm='-', Vsd='-', P=0, r='-', flag='F')  # Finished
        wx.PostEvent(self.RunPage, stop_ev)
        stat_ev = evts.StatusEvent(msg='RLINK RUN COMPLETED', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

        self.RunPage.RLinkBtn.Enable(True)
        self.RunPage.StartBtn.Enable(True)
        self.RunPage.StopBtn.Enable(False)

    def standby(self):
        self.RunPage.V1Setting.SetValue('0')
        self.RunPage.V2Setting.SetValue('0')

    def abort(self):
        """abort worker thread."""
        # Method for use by main thread to signal an abort
        stat_ev = evts.StatusEvent(msg='abort(): Run aborted', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        self._want_abort = 1

    # def get_multiplier(self, name):
    #     # A helper function to extract the value multiplier from resistor name
    #     mult = 0
    #     if name[-1] == 'k':
    #         mult = 1000
    #     elif name[-1] == 'M':
    #         mult = int(1e6)
    #     elif name[-1] == 'G':
    #         mult = int(1e9)
    #     return mult

    # def filt(self, char):
    #     # A helper function to strip rubbish from DVM o/p unicode string
    #     # and retain any number (as a string)
    #     accept_str = u'-0.12345678eE9'
    #     return char in accept_str  # Returns 'True' or 'False'

"""--------------End of Thread class definition-------------------"""
